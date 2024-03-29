import openpyxl
import numpy as np
from processing import Processing

class Utils:
    dataset_path = './data/motions/'
    processing = Processing()
    def __init__(self, xlxs_path) -> None:
        self.xlxs_path = xlxs_path

    def load_xlsx(self):
        wb = openpyxl.load_workbook(self.xlxs_path)
        self.ws = wb.active
        return self.ws['A1'].value
        
    def combine_selected_motion(self):
        is_first_data = True
        motions_len = []
        interpo_len = []
        for i, row in enumerate(self.ws.iter_rows(values_only=True, min_row=2)):
            anim = row[0]
            interpo_len.append(row[1])
            if anim is not None:
                data = np.load(self.dataset_path + anim)
                if is_first_data != True:
                    motions = np.concatenate((motions,data))
                else:
                    motions = data
                    is_first_data = False
                motions_len.append(len(data))
        assert len(motions_len) >= 2
        motions = self.processing.normalize(motions)
        motions = self.processing.calculate_angle(motions)
        interpo_len[-1] = 0
        return motions, motions_len, interpo_len
    
if __name__ == "__main__":
    file_name = "template_frames_160_bpm_90.xlsx"
    utils = Utils(file_name)
    utils.load_xlsx()
    data, _, _ = utils.combine_selected_motion()
    np.save("test.npy",data)
